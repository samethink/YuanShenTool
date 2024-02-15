"""
Author: ithink
Create: 2024.2.13 12:10
Project: YuanShenTool
Path: src/modules/inv.py
IDE: PyCharm
Description: 获取/操作需求清单
"""
import os

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, PatternFill

from src.utils.cyber import UA
from src.utils.support import SYSTEM_NAME, logger, pub_config

if not os.path.exists('cache/'):
    os.makedirs('cache/')


class Web:
    """网络接口的请求"""

    def __init__(self):
        self.__cookie_path = 'cache/mys_cookie.txt'
        if os.path.exists(self.__cookie_path):
            with open(self.__cookie_path, 'r', encoding='ASCII') as fp:
                self.cookie = fp.read()
        else:
            self.cookie = None

    def validate_cookie(self):
        return self.cookie and len(self.cookie) > 50 and self.cookie.isascii()

    def set_cookie(self, value):
        """更新cookie值并保存

        :param value: 新值
        :return:
        """
        self.cookie = value
        logger.debug('cookie: %s' % self.cookie)
        if self.validate_cookie():
            with open(self.__cookie_path, 'w', encoding='ASCII') as fp:
                fp.write(self.cookie)

    def get_inventory(self, share_code):
        """摹本物品列表的请求方法

        :param share_code: 摹本分享码
        :return: 接口返回数据
        """
        url = 'https://api-takumi.mihoyo.com/event/e20200928calculate/v1/furniture/blueprint'
        params = {
            'share_code': share_code,
            'region': 'cn_gf01'
        }
        headers = {
            'Cookie': self.cookie,
            'Referer': 'https://webstatic.mihoyo.com/',
            'User-Agent': UA
        }
        logger.debug('==> GET %s P=%s' % (url, params))
        response = requests.get(url, params=params, headers=headers)
        logger.debug('<== %s %s %s' % (response.status_code, url, response.text[:100]))
        return response.json()


class FetchInv:
    """获取需求清单"""

    def __init__(self):
        self.web = Web()
        self.__workbook = Workbook()
        self.__worksheet = self.__workbook.active

        self.__icon_dir = 'cache/item_icons/'
        if os.path.exists(self.__icon_dir):
            self.__item_ids = {_[:-4] for _ in os.listdir(self.__icon_dir)}
        else:
            self.__item_ids = {}
            os.makedirs(self.__icon_dir)

    def download_inventory(self, share_code):
        """请求接口获得列表后，提取数据并保存为Excel文件

        :param share_code: 分享码
        :return: 成功标志，消息
        """
        if not share_code:
            return False, '摹本分享码不能为空'
        elif not self.web.validate_cookie():
            return False, '<set_cookie>'

        try:
            result = self.web.get_inventory(share_code)
            if result['data']:
                info_list = result['data']['list'] + result['data']['not_calc_list']
                filename = 'inventory_%s.xlsx' % share_code
                self.save_inventory_as_xlsx(info_list, filename)
                return True, filename
            elif result['retcode'] == -100:
                self.web.set_cookie(value='')
                return False, 'cookie已失效：' + result['message']
            else:
                return False, '失败：' + result['message']
        except requests.RequestException:
            return False, '无网络连接'
        except requests.JSONDecodeError:
            return False, '响应结果解析失败'

    def save_inventory_as_xlsx(self, data: list[dict], filename):
        """保存文件

        :param data: 响应数据的物品列表
        :param filename: 保存文件名
        :return:
        """
        saved_path = 'cache/' + filename
        data.sort(key=lambda x: x['num'], reverse=True)

        titles = ['ID', '等级', '名称', '图片', '需求数量', '已有数量']
        self.__worksheet.append(titles)
        for col in range(1, 7):
            cell = self.__worksheet.cell(1, col)
            cell.fill = PatternFill(start_color='00DDBB', fill_type='solid')
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        self.__worksheet.row_dimensions[1].height = 20

        for i in range(len(data)):
            row = i + 2
            self.__worksheet.cell(row, 1).value = '=HYPERLINK("%s", %s)' % (data[i]['wiki_url'], data[i]['id']) \
                if data[i]['wiki_url'] \
                else str(data[i]['id'])
            self.__worksheet.cell(row, 2).value = data[i]['level']
            self.__worksheet.cell(row, 3).value = data[i]['name']

            if pub_config['insert_image']:
                try:
                    image = self.get_item_icon(data[i])
                    image.width, image.height = 40, 40
                    self.__worksheet.add_image(image, 'D' + str(row))
                except requests.RequestException:
                    self.__worksheet.cell(row, 4).value = '!err'
            else:
                self.__worksheet.cell(row, 4).value = 'NoImg'

            self.__worksheet.cell(row, 5).value = data[i]['num']

            self.__worksheet.row_dimensions[row].height = 30
            for col in range(1, 7):
                self.__worksheet.cell(row, col).alignment = Alignment(horizontal='center', vertical='center')

        self.__worksheet.column_dimensions['C'].width = 40
        self.__worksheet.column_dimensions['D'].width = 5
        self.__workbook.save(saved_path)
        os.system('start ' if SYSTEM_NAME == 'Windows' else 'open ' + saved_path)

    def get_item_icon(self, item):
        """获取物品的图标，本地无对应图标时下载并缓存

        :param item: 保存物品信息的项目
        :return: Image对象
        """
        icon_path = '%s/%s.png' % (self.__icon_dir, item['id'])
        if str(item['id']) not in self.__item_ids:
            with open(icon_path, 'wb') as fp:
                fp.write(requests.get(
                    item['icon_url'],
                    headers={'Connection': 'keep-alive', 'User-Agent': UA}
                ).content)
                logger.debug('icon[%s]已缓存' % item['id'])
        return Image(icon_path)


class HandleInv:
    """使用本地清单数据

    初始化会读取指定Excel文件的数据转为字典，缓存到data属性
    保存时会将更改后的data数据写入到源文件中
    """

    def __init__(self, xlsx_filename):
        self.__source_path = 'cache/' + xlsx_filename
        self.__workbook = load_workbook(self.__source_path)
        self.__worksheet = self.__workbook.active
        self.data = {}
        self._set_data()

    def _set_data(self):
        """读取Excel表数据"""
        for row in range(2, self.__worksheet.max_row + 1):
            self.data[self.__worksheet.cell(row, 3).value] = [
                self.__worksheet.cell(row, 5).value,
                0 if (_ := self.__worksheet.cell(row, 6).value) is None else _
            ]

    def save_data(self):
        """将data的数据保存到源文件，并关闭工作簿"""
        for row in range(2, self.__worksheet.max_row + 1):
            key = self.__worksheet.cell(row, 3).value
            self.__worksheet.cell(row, 5).value = self.data[key][0]
            self.__worksheet.cell(row, 6).value = self.data[key][1]
        self.__workbook.save(self.__source_path)
        self.__workbook.close()


def get_inv_filelist():
    return [filename for filename in os.listdir('cache/') if filename.startswith('inventory')]


if __name__ == '__main__':
    inventory = HandleInv('inventory_4516178075.xlsx')
    print(inventory.data)
    inventory.data['「桔梗执别愁云去」'][0] += 1
    inventory.data['「桔梗执别愁云去」'][1] -= 100
    print(inventory.data)
    inventory.save_data()
